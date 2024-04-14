import json
from openpyxl.reader.excel import load_workbook


def load_wb(f_path):
    workbook = load_workbook(f_path)
    datas = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        title = sheet[1]
        for index in range(2, sheet.max_row + 1):
            row = sheet[index]
            data = {}
            for col_index in range(len(row)):
                data[title[col_index].value] = row[col_index].value
            datas.append(data)
    return datas


def read_json(path):
    json_list = []
    with open(path, encoding="utf-8") as f:
        for i in f:
            json_list.append(json.loads(i))
    return json_list
