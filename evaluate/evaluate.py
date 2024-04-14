import json

import torch
from openpyxl.reader.excel import load_workbook
from transformers import AutoTokenizer, AutoModelForCausalLM
from openxlab.model import download

from uitls.logger import get_logger

model_name = "/home/ssc/model/ILM-7B-SFT/"

tokenizer = AutoTokenizer.from_pretrained(model_name, trust_remote_code=True)
model = AutoModelForCausalLM.from_pretrained(model_name, device_map="auto", trust_remote_code=True, torch_dtype=torch.float16)

model = model.eval()

datas = []

template = "多选题：{content} \n选项A：{A} \n选项B：{B} \n选项C：{C} \n选项D：{D} \n请选择正确的一个或多个答案。"


def evaluate(f_path, logger, is_subjective=False, has_answer=False):
    workbook = load_workbook(f_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        with open("./qa_chat.jsonl", mode="w+", encoding="utf-8") as file:
            for index in range(2, sheet.max_row + 1):
                row = sheet[index]
                if is_subjective:
                    prompt = row[1].value
                else:
                    prompt = template.format(content=row[0].value,
                                             A=row[1].value,
                                             B=row[2].value,
                                             C=row[3].value,
                                             D=row[4].value)

                response, history = model.chat(tokenizer, prompt, history=[], temperature=0.1)

                if has_answer:
                    true_answer = row[2].value
                    logger.info("Prompt：" +
                                prompt.replace("\n", "\n                                        ") +
                                "\n                                        LLM Response：" +
                                response +
                                "\n                                        True Answer："+
                                true_answer)
                else:
                    logger.info("Prompt：" +
                                prompt.replace("\n", "\n                                        ") +
                                "\n                                        LLM Response：" +
                                response)



# f_path = "./data/赛道二-主观题.xlsx"
# logger = get_logger("./subjective.txt")
# evaluate(f_path, logger, is_subjective=True)

f_path = "./data/赛道二-测试题.xlsx"
logger = get_logger("./objective.txt")
evaluate(f_path, logger)